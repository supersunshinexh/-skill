#!/usr/bin/env python3
"""Compare reimbursement totals between old and rebuilt source tables.

Use this before HuiLianYi browser entry when rebuilding reports from invoices,
screenshots, or workbooks. It highlights total differences so missing payment
proof rows are caught before a smaller report is created.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


AMOUNT_ALIASES = ["金额", "报销金额", "付款金额", "实付金额", "价税合计", "汇联易单据金额"]
DEFAULT_GROUP_COLUMNS = ["报销单类型", "发票抬头", "采购原因"]


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def amount(value: Any) -> Decimal:
    text = clean(value)
    if not text:
        return Decimal("0.00")
    text = text.replace(",", "").replace("￥", "").replace("¥", "").replace("CNY", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return Decimal("0.00")
    try:
        return Decimal(match.group(0)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [dict(row) for row in csv.DictReader(fh)]


def first_nonempty_row(rows: list[tuple[Any, ...]]) -> int | None:
    for index, row in enumerate(rows):
        if any(clean(cell) for cell in row):
            return index
    return None


def read_xlsx(path: Path, sheets: list[str] | None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet_names = sheets or workbook.sheetnames
    out: list[dict[str, Any]] = []

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        header_index = first_nonempty_row(rows)
        if header_index is None:
            continue
        headers = [clean(cell) for cell in rows[header_index]]
        if not any(header in AMOUNT_ALIASES for header in headers):
            continue
        for raw in rows[header_index + 1 :]:
            if not any(clean(cell) for cell in raw):
                continue
            row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw))) if headers[i]}
            row["__sheet"] = sheet_name
            out.append(row)
    return out


def read_table(path: Path, sheets: list[str] | None = None) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path, sheets)
    raise ValueError(f"Unsupported file type: {path}")


def pick_amount_column(row: dict[str, Any]) -> str | None:
    for name in AMOUNT_ALIASES:
        if name in row and clean(row[name]):
            return name
    return None


def group_key(row: dict[str, Any], columns: list[str]) -> str:
    values = []
    for column in columns:
        value = clean(row.get(column, ""))
        if column == "发票抬头" and value == "无票":
            value = "凭证费用"
        values.append(value or "(空)")
    return " / ".join(values)


def summarize(rows: list[dict[str, Any]], group_columns: list[str]) -> dict[str, dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = defaultdict(lambda: {"amount": Decimal("0.00"), "rows": 0, "items": []})
    for row in rows:
        amount_column = pick_amount_column(row)
        if not amount_column:
            continue
        key = group_key(row, group_columns)
        value = amount(row[amount_column])
        groups[key]["amount"] += value
        groups[key]["rows"] += 1
        groups[key]["items"].append(
            {
                "sheet": clean(row.get("__sheet")),
                "name": clean(row.get("费用名称") or row.get("资产名称") or row.get("项目名称")),
                "amount": str(value),
                "reason": clean(row.get("采购原因") or row.get("事由") or row.get("备注")),
            }
        )
    return groups


def compare(old_groups: dict[str, dict[str, Any]], new_groups: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for key in sorted(set(old_groups) | set(new_groups)):
        old_amount = old_groups.get(key, {}).get("amount", Decimal("0.00"))
        new_amount = new_groups.get(key, {}).get("amount", Decimal("0.00"))
        delta = new_amount - old_amount
        if delta == Decimal("0.00"):
            continue
        out.append(
            {
                "group": key,
                "old_amount": str(old_amount.quantize(Decimal("0.01"))),
                "new_amount": str(new_amount.quantize(Decimal("0.01"))),
                "delta": str(delta.quantize(Decimal("0.01"))),
                "old_rows": old_groups.get(key, {}).get("rows", 0),
                "new_rows": new_groups.get(key, {}).get("rows", 0),
                "old_items": old_groups.get(key, {}).get("items", []),
                "new_items": new_groups.get(key, {}).get("items", []),
            }
        )
    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--old", required=True, type=Path, help="Prior validated workbook/export")
    parser.add_argument("--new", required=True, type=Path, help="Rebuilt workbook/export")
    parser.add_argument(
        "--group-columns",
        default=",".join(DEFAULT_GROUP_COLUMNS),
        help="Comma-separated grouping columns. Defaults to report type, invoice title, purchase reason.",
    )
    parser.add_argument("--old-sheet", action="append", help="Old workbook sheet to read. Repeat for multiple sheets.")
    parser.add_argument("--new-sheet", action="append", help="New workbook sheet to read. Repeat for multiple sheets.")
    parser.add_argument("--fail-on-diff", action="store_true", help="Exit 2 when any total difference is found")
    args = parser.parse_args()

    group_columns = [column.strip() for column in args.group_columns.split(",") if column.strip()]
    old_groups = summarize(read_table(args.old, args.old_sheet), group_columns)
    new_groups = summarize(read_table(args.new, args.new_sheet), group_columns)
    differences = compare(old_groups, new_groups)

    result = {
        "old": str(args.old),
        "new": str(args.new),
        "group_columns": group_columns,
        "difference_count": len(differences),
        "differences": differences,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 2 if differences and args.fail_on_diff else 0


if __name__ == "__main__":
    raise SystemExit(main())
