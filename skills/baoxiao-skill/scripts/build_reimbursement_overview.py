#!/usr/bin/env python3
"""Build a HuiLianYi reimbursement overview workbook.

Inputs are extracted evidence rows and, optionally, fixed-asset rows. The script
does not perform OCR; use it after invoice PDFs and no-ticket screenshots have
been extracted into CSV/XLSX tables.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT_COLUMNS = [
    "发票抬头",
    "费用名称",
    "资产编号",
    "使用人",
    "使用部门",
    "金额",
    "汇联易报销单号",
    "采购原因",
    "报销单类型",
    "汇联易单据金额",
    "报销备注",
]

ALIASES = {
    "发票抬头": ["发票抬头", "票据抬头", "抬头", "购买方", "发票购买方"],
    "费用名称": ["费用名称", "资产名称", "物品名称", "项目名称", "商品名称", "名称"],
    "资产编号": ["资产编号", "固定资产编号", "资产编码", "编号"],
    "使用人": ["使用人", "领用人", "员工", "申请人"],
    "使用部门": ["使用部门", "部门", "归属部门"],
    "金额": ["金额", "价税合计", "付款金额", "实付金额", "小计"],
    "汇联易报销单号": ["汇联易报销单号", "报销单号", "单号", "汇联易单号"],
    "采购原因": ["采购原因", "用途", "备注用途", "事由", "付款事由"],
    "报销单类型": ["报销单类型", "单据类型", "汇联易单据类型"],
    "汇联易单据金额": ["汇联易单据金额", "报销单金额", "单据金额", "汇联易金额"],
    "报销备注": ["报销备注", "备注", "核对备注"],
}

TEXT_MATCH_COLUMNS = ["费用名称", "采购原因", "使用部门", "发票抬头"]
MERGE_COLUMNS = ["汇联易报销单号", "汇联易单据金额"]


@dataclass
class MatchResult:
    row: dict[str, Any] | None
    score: float


def read_table(path: Path, sheet: str | None = None) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            return [dict(row) for row in csv.DictReader(fh)]

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    header_index = next(
        (i for i, row in enumerate(rows) if any(cell not in (None, "") for cell in row)),
        None,
    )
    if header_index is None:
        return []

    headers = [clean_text(cell) for cell in rows[header_index]]
    out: list[dict[str, Any]] = []
    for raw_row in rows[header_index + 1 :]:
        if not any(cell not in (None, "") for cell in raw_row):
            continue
        out.append({headers[i]: raw_row[i] for i in range(min(len(headers), len(raw_row)))})
    return out


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def first_value(row: dict[str, Any], column: str) -> Any:
    for name in ALIASES[column]:
        if name in row and clean_text(row[name]) != "":
            return row[name]
    return ""


def canonicalize(row: dict[str, Any]) -> dict[str, Any]:
    out = {column: first_value(row, column) for column in OUTPUT_COLUMNS}
    out["金额"] = amount_decimal(out["金额"])
    out["汇联易单据金额"] = amount_decimal(out["汇联易单据金额"])

    title = clean_text(out["发票抬头"])
    doc_type = clean_text(out["报销单类型"])
    row_text = " ".join(clean_text(value) for value in row.values())
    no_ticket = re.search(r"无票|截图|支付凭证|付款截图|支付截图", f"{title} {doc_type} {row_text}")
    if no_ticket:
        out["发票抬头"] = "无票"
    elif not title:
        out["发票抬头"] = "发票抬头待核对"
        note = clean_text(out["报销备注"])
        out["报销备注"] = "发票抬头待核对" if not note else f"{note}; 发票抬头待核对"
    if not doc_type:
        out["报销单类型"] = "无票报销单" if out["发票抬头"] == "无票" else "日常报销单"
    return out


def amount_decimal(value: Any) -> Decimal | str:
    text = clean_text(value)
    if not text:
        return ""
    text = (
        text.replace(",", "")
        .replace("￥", "")
        .replace("¥", "")
        .replace("CNY", "")
        .replace("人民币", "")
        .strip()
    )
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return clean_text(value)
    try:
        return Decimal(match.group(0)).quantize(Decimal("0.01"))
    except InvalidOperation:
        return clean_text(value)


def amount_equal(left: Any, right: Any) -> bool:
    if not isinstance(left, Decimal) or not isinstance(right, Decimal):
        return False
    return abs(left - right) <= Decimal("0.01")


def text_score(left: Any, right: Any) -> float:
    a = re.sub(r"\s+", "", clean_text(left))
    b = re.sub(r"\s+", "", clean_text(right))
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return 0.85
    return SequenceMatcher(None, a, b).ratio()


def best_asset_match(evidence: dict[str, Any], assets: list[dict[str, Any]]) -> MatchResult:
    best = MatchResult(None, 0.0)
    for asset in assets:
        score = 0.0
        if amount_equal(evidence["金额"], asset["金额"]):
            score += 50.0
        elif isinstance(evidence["金额"], Decimal) and isinstance(asset["金额"], Decimal):
            continue

        for column in TEXT_MATCH_COLUMNS:
            weight = 30.0 if column == "费用名称" else 10.0
            score += text_score(evidence[column], asset[column]) * weight

        if clean_text(evidence["资产编号"]) and clean_text(evidence["资产编号"]) == clean_text(asset["资产编号"]):
            score += 80.0

        if score > best.score:
            best = MatchResult(asset, score)
    return best


def enrich_with_assets(evidence_rows: list[dict[str, Any]], asset_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not asset_rows:
        return evidence_rows

    assets = [canonicalize(row) for row in asset_rows]
    enriched: list[dict[str, Any]] = []
    for row in evidence_rows:
        if all(clean_text(row[column]) for column in ["资产编号", "使用人", "使用部门"]):
            enriched.append(row)
            continue

        match = best_asset_match(row, assets)
        if match.row and match.score >= 60:
            merged = dict(row)
            for column in ["费用名称", "资产编号", "使用人", "使用部门", "采购原因"]:
                if clean_text(merged[column]) == "":
                    merged[column] = match.row[column]
            if clean_text(merged["报销备注"]) == "":
                merged["报销备注"] = "资产信息已匹配"
            enriched.append(merged)
        else:
            pending = dict(row)
            note = clean_text(pending["报销备注"])
            pending["报销备注"] = "资产匹配待核对" if not note else f"{note}; 资产匹配待核对"
            enriched.append(pending)
    return enriched


def fill_report_totals(rows: list[dict[str, Any]]) -> None:
    totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0.00"))
    for row in rows:
        report_id = clean_text(row["汇联易报销单号"])
        if report_id and isinstance(row["金额"], Decimal):
            totals[report_id] += row["金额"]

    for row in rows:
        report_id = clean_text(row["汇联易报销单号"])
        if report_id and row["汇联易单据金额"] == "":
            row["汇联易单据金额"] = totals[report_id].quantize(Decimal("0.01"))


def write_workbook(rows: list[dict[str, Any]], output: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "报销一览表"

    blue_fill = PatternFill("solid", fgColor="1F4E78")
    green_fill = PatternFill("solid", fgColor="92D050")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    worksheet.append(OUTPUT_COLUMNS)
    blue_headers = {"汇联易报销单号", "报销单类型", "汇联易单据金额", "报销备注"}
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF" if cell.value in blue_headers else "000000")
        cell.fill = blue_fill if cell.value in blue_headers else white_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in rows:
        worksheet.append([row[column] for column in OUTPUT_COLUMNS])

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if worksheet.cell(1, cell.column).value == "金额":
                cell.fill = green_fill
                if isinstance(cell.value, Decimal):
                    cell.value = float(cell.value)
            elif isinstance(cell.value, Decimal):
                cell.value = float(cell.value)

    widths = {
        "A": 22,
        "B": 26,
        "C": 24,
        "D": 12,
        "E": 42,
        "F": 12,
        "G": 20,
        "H": 34,
        "I": 16,
        "J": 18,
        "K": 30,
    }
    for col, width in widths.items():
        worksheet.column_dimensions[col].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    merge_repeated_cells(worksheet, MERGE_COLUMNS)

    for column in ["金额", "汇联易单据金额"]:
        idx = OUTPUT_COLUMNS.index(column) + 1
        for cell in worksheet.iter_cols(min_col=idx, max_col=idx, min_row=2):
            for item in cell:
                item.number_format = '#,##0.00'

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def merge_repeated_cells(worksheet, columns: list[str]) -> None:
    if worksheet.max_row < 2:
        return
    for column in columns:
        index = OUTPUT_COLUMNS.index(column) + 1
        start = 2
        previous = worksheet.cell(start, index).value
        for row_num in range(3, worksheet.max_row + 2):
            current = worksheet.cell(row_num, index).value if row_num <= worksheet.max_row else None
            if current != previous:
                if previous not in (None, "") and row_num - start > 1:
                    worksheet.merge_cells(
                        start_row=start,
                        start_column=index,
                        end_row=row_num - 1,
                        end_column=index,
                    )
                    worksheet.cell(start, index).alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                start = row_num
                previous = current


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--evidence", required=True, type=Path, help="CSV/XLSX evidence rows")
    parser.add_argument("--assets", type=Path, help="Optional CSV/XLSX fixed-asset rows")
    parser.add_argument("--evidence-sheet", help="Sheet name for evidence workbook")
    parser.add_argument("--assets-sheet", help="Sheet name for asset workbook")
    parser.add_argument("--output", required=True, type=Path, help="Output XLSX path")
    args = parser.parse_args()

    evidence_rows = [canonicalize(row) for row in read_table(args.evidence, args.evidence_sheet)]
    asset_rows = read_table(args.assets, args.assets_sheet) if args.assets else []
    rows = enrich_with_assets(evidence_rows, asset_rows)
    fill_report_totals(rows)
    write_workbook(rows, args.output)
    print(f"wrote {args.output} ({len(rows)} rows)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
